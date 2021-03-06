import numpy as np
import scipy.io as sio
import os

from pacu.core.io.scanbox.view.trial_merged_roi import TrialMergedROIView

from cStringIO import StringIO
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font

class Export(object):
    '''
    Create this object in directory that contains the .io files of interest.
    Argument should be list of tuples with the filename first and the workspace name second.
    Example: stitched_dataset = stitched_data([('Day1_000_000','Workspace_1'),('Day1_000_001',Workspace_1')])
    '''
    def __init__(self, condition, ids, rois): #, fw_array):
        #self.fw_array = fw_array
        #self.path = os.getcwd()
        #self.io = [ScanboxIO(os.path.join(self.path,fw[0])) for fw in self.fw_array]
        #self.io = io
        #self.wsName = wsName
        self.condition = condition
        self.rois = rois
        self.ids = map(int, ids.split(','))
        print(self.ids)
        #self.active_workspace = io.condition.workspaces.filter_by(name = self.wsName)[0]
        #self.condition = self.io.condition
        #self.workspaces = [workspace for data,fw in zip(self.io, self.fw_array) for workspace in data.condition.workspaces if workspace.name == fw[1]]
        #self.workspaces = self.io.workspaces
        #self.rois = self.find_matched_rois()
        #self.merged_rois = [TrialMergedROIView(roi.params.cell_id,*self.workspaces) for roi in self.rois[0]]
        #self.refresh_all() # no need to refresh unless stitching the data
        #self.roi_dict = {'{}{}'.format('cell_id_',merged_roi.rois[0].params.cell_id):merged_roi.serialize() for merged_roi in self.merged_rois}

    def find_matched_rois(self):
        rois = [workspace.rois for data, fw in zip(self.io, self.fw_array) for workspace in data.condition.workspaces if workspace.name == fw[1]]
        id_sets = [[roi.params.cell_id for roi in roi_list] for roi_list in rois]
        list_lengths = [len(s) for s in id_sets]
        shortest_idx = list_lengths.index(min(list_lengths))
        shortest_set = id_sets.pop(shortest_idx)
        matched_ids = set(shortest_set).intersection(*id_sets)
        matched_rois = [[roi for roi in roi_list if roi.params.cell_id in matched_ids] for roi_list in rois]
        return matched_rois

    def refresh_all(self):
        for merged_roi in self.merged_rois:
            merged_roi.refresh()

    def sorted_orientation_traces(self, merged_roi):
        sorted_orientation_traces = {str(roi.workspace.name):{} for roi in merged_roi.rois}
        for k in sorted_orientation_traces.keys():
            sorted_orientation_traces[k] = [dict(dtorientationsmean.attributes.items()[i] for i in [8,17]) for roi in merged_roi.rois for dtorientationsmean in roi.dtorientationsmeans if roi.workspace.name == k]
        return sorted_orientation_traces

    def blank_responses(self, merged_roi):
        blank_responses = [trial.attributes['value']['on']
                            for trial in merged_roi.dff0s
                            if trial.attributes['trial_blank'] == True]
        return np.mean(blank_responses, 1)

    def flicker_responses(self, merged_roi):
        flicker_responses = [trial.attributes['value']['on']
                            for trial in merged_roi.dff0s
                            if trial.attributes['trial_flicker'] == True]
        return np.mean(flicker_responses, 1)

    def export_mat(self,filename=None,p_value=.01,unmerge=0):
        merged_dict = {}
        merged_dict['filenames'] = [fw[0][:-3] for fw in self.fw_array]
        merged_dict['workspaces'] = [workspace.name for workspace in self.workspaces]
        merged_dict['rois'] = self.roi_dict
        for merged_roi in self.merged_rois:
            merged_dict['rois']['{}{}'.format('cell_id_',merged_roi.rois[0].params.cell_id)]['sorted_dtorientationsmeans'] = self.sorted_orientation_traces(merged_roi)
            merged_dict['rois']['{}{}'.format('cell_id_',merged_roi.rois[0].params.cell_id)]['blank_responses'] = self.blank_responses(merged_roi)
            merged_dict['rois']['{}{}'.format('cell_id_',merged_roi.rois[0].params.cell_id)]['flicker_responses'] = self.flicker_responses(merged_roi)
        if filename == None:
            fname = self.fw_array[0][0][:-3] + '_merged.mat'
        else:
            fname = filename + '.mat'
        # save .mat file
        sio.savemat(fname,{'merged_dict':merged_dict})

    def excel(self, p_value=0.01, unmerge=0):
        # export data as .xlsx
        wb = Workbook()
        ws = wb.active

        filtered_rois = [roi for roi in self.rois if roi.id in self.ids]
        #num_rois = len(self.rois)
        num_rois = len(filtered_rois)
        sfreqs = self.condition.sfrequencies #attributes['sfrequencies']
        num_sf = len(sfreqs)
        idx_list = range(3, num_rois*num_sf, num_sf)
        font = 'Courier New'

        # format header columns
        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:C1')
        ws.merge_cells('D1:E1')
        ws.merge_cells('F1:I1')
        ws.merge_cells('S1:T1')

        for top,bottom in zip(ws['J1:Q1'][0],ws['J2:Q2'][0]):
            ws.merge_cells('{}{}:{}{}'.format(top.column,top.row,bottom.column,bottom.row))

        header = NamedStyle(name='header')
        header.alignment = Alignment(horizontal='center',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=True,
                                        shrink_to_fit=False,
                                        indent=0)

        header.font = Font(name=font,
                                size=10,
                                bold=True,
                                italic=False,
                                vertAlign=None,
                                underline='none',
                                strike=False,
                                color='FF000000')

        header.border = Border(top=Side(border_style='medium',
                                        color='FF000000'),
                                    bottom=Side(border_style='medium',
                                        color='FF000000')
                                    )

        reg_cell = NamedStyle(name='regular')
        reg_cell.alignment = Alignment(horizontal=None,
                                     vertical='center',
                                     text_rotation=0,
                                     wrap_text=True,
                                     shrink_to_fit=False,
                                     indent=0)

        reg_cell.font = Font(name=font,size=10)

        reg_cell.border = Border(top=Side(border_style='medium',
                                                 color='FF000000'),
                                             bottom=Side(border_style='medium',
                                                 color='FF000000')
                                             )

        sig_cell = NamedStyle(name='significant')
        sig_cell.alignment = Alignment(horizontal=None,
                                          vertical='center',
                                          text_rotation=0,
                                          wrap_text=True,
                                          shrink_to_fit=False,
                                          indent=0)

        sig_cell.font = Font(name=font,size=10)

        sig_cell.border = Border(top=Side(border_style='medium',
                                                 color='FF000000'),
                                             bottom=Side(border_style='medium',
                                                 color='FF000000')
                                             )

        sig_cell.fill = PatternFill(start_color='FFFFFF00',
                                        end_color='FFFFFF00',
                                        fill_type='solid')

        # write column titles
        ws['A1'].value = 'Cell ID'
        ws['A1'].style = header
        ws['B1'].value = 'Anova All'
        ws['B1'].style = header
        ws['D1'].value = 'SF Cutoff Rel33'
        ws['D1'].style = header
        ws['F1'].value = 'SF'
        ws['F1'].style = header
        ws['S1'].value = 'Anova Each'
        ws['S1'].style = header
        ws['S2'].value = 'F'
        ws['S2'].style = header
        ws['T2'].value = 'P'
        ws['T2'].style = header

        for cell,val in zip(ws[2][1:9], ['F','P','X','Y','Peak','Pref','Bandwidth','Global\nOPref']):
            cell.value = val
            cell.style = header

        for cell,val in zip(ws[1][9:18], ['@', 'OSI', 'CV', 'DCV', 'DSI', 'Sigma', 'OPref', 'RMax', 'Residual']):
            cell.value = val
            cell.style = header

        for idx,roi in zip(idx_list, filtered_rois):
            print(roi.id)
            peak_sf = round(roi.dtsfreqfits[0].attributes['value']['peak'],2)
            try:
                if roi.dtanovaeachs.filter_by(trial_sf=peak_sf)[0].p <= p_value:
                    style = sig_cell
                else:
                    style = reg_cell
            except IndexError:
                style = reg_cell

            if unmerge == 0:
                for top,bottom in zip(ws['A{}:I{}'.format(idx, idx)][0], ws['A{}:I{}'.format(idx+num_sf-1, idx+num_sf-1)][0]):
                    ws.merge_cells('{}{}:{}{}'.format(top.column, top.row, bottom.column, bottom.row))

                ws.cell(row=idx,column=1).value = int(roi.params.cell_id)
                ws.cell(row=idx,column=1).style = style
                ws.cell(row=idx,column=2).value = roi.dtanovaalls.first.attributes['value']['f']
                ws.cell(row=idx,column=2).style = style
                ws.cell(row=idx,column=3).value = roi.dtanovaalls.first.attributes['value']['p']
                ws.cell(row=idx,column=3).style = style
                try:
                    ws.cell(row=idx,column=4).value = roi.dtsfreqfits.first.attributes['value']['rc33'].x
                    ws.cell(row=idx,column=5).value = roi.dtsfreqfits.first.attributes['value']['rc33'].y
                except AttributeError:
                    #print 'No "SF Cutoff Rel33" found for cell ',roi.params.cell_id
                    ws.cell(row=idx,column=4).value = None
                    ws.cell(row=idx,column=5).value = None
                ws.cell(row=idx,column=4).style = style
                ws.cell(row=idx,column=5).style = style
                ws.cell(row=idx,column=6).value = peak_sf
                ws.cell(row=idx,column=6).style = style
                ws.cell(row=idx,column=7).value = roi.dtsfreqfits.first.attributes['value']['pref']
                ws.cell(row=idx,column=7).style = style
                ws.cell(row=idx,column=8).value = roi.dtsfreqfits.first.attributes['value']['ratio']
                ws.cell(row=idx,column=8).style = style
                ws.cell(row=idx,column=9).value = roi.dtorientationbestprefs.first.attributes['value']
                ws.cell(row=idx,column=9).style = style

            elif unmerge == 1:
                for i in range(num_sf):
                    ws.cell(row=idx+i,column=1).value = int(roi.params.cell_id)
                    ws.cell(row=idx+i,column=1).style = style
                    ws.cell(row=idx+i,column=2).value = roi.dtanovaalls.first.attributes['value']['f']
                    ws.cell(row=idx+i,column=2).style = style
                    ws.cell(row=idx+i,column=3).value = roi.dtanovaalls.first.attributes['value']['p']
                    ws.cell(row=idx+i,column=3).style = style
                    try:
                        ws.cell(row=idx+i,column=4).value = roi.dtsfreqfits.first.attributes['value']['rc33'].x
                        ws.cell(row=idx+i,column=5).value = roi.dtsfreqfits.first.attributes['value']['rc33'].y
                    except AttributeError:
                        #print 'No "SF Cutoff Rel33" found for cell ',roi.params.cell_id
                        ws.cell(row=idx+i,column=4).value = None
                        ws.cell(row=idx+i,column=5).value = None
                    ws.cell(row=idx+i,column=4).style = style
                    ws.cell(row=idx+i,column=5).style = style
                    ws.cell(row=idx+i,column=6).value = peak_sf
                    ws.cell(row=idx+i,column=6).style = style
                    ws.cell(row=idx+i,column=7).value = roi.dtsfreqfits.first.attributes['value']['pref']
                    ws.cell(row=idx+i,column=7).style = style
                    ws.cell(row=idx+i,column=8).value = roi.dtsfreqfits.first.attributes['value']['ratio']
                    ws.cell(row=idx+i,column=8).style = style
                    ws.cell(row=idx+i,column=9).value = roi.dtorientationbestprefs.first.attributes['value']
                    ws.cell(row=idx+i,column=9).style = style

            for i,cell in enumerate(ws.iter_rows(min_col=10, max_col=10, min_row=idx, max_row=idx+num_sf-1)):
                    cell[0].value = sfreqs[i]
                    cell[0].style = style

            for i,row in enumerate(ws.iter_rows(min_col=11, max_col=20, min_row=idx, max_row=idx+num_sf-1)):

                row[0].value = roi.dtorientationsfits[i].attributes['value']['osi']
                row[0].style = style
                row[1].value = roi.dtorientationsfits[i].attributes['value']['cv']
                row[1].style = style
                row[2].value = roi.dtorientationsfits[i].attributes['value']['dcv']
                row[2].style = style
                row[3].value = roi.dtorientationsfits[i].attributes['value']['dsi']
                row[3].style = style
                row[4].value = roi.dtorientationsfits[i].attributes['value']['sigma']
                row[4].style = style
                row[5].value = roi.dtorientationsfits[i].attributes['value']['o_pref']
                row[5].style = style
                row[6].value = roi.dtorientationsfits[i].attributes['value']['r_max']
                row[6].style = style
                row[7].value = roi.dtorientationsfits[i].attributes['value']['residual']
                row[7].style = style
                row[8].value = roi.dtanovaeachs[i].attributes['f']
                row[8].style = style
                row[9].value = roi.dtanovaeachs[i].attributes['p']
                row[9].style = style

        # save excel file
        #if filename == None:
        #    fname = self.fw_array[0][0][:-3] + '_merged.xlsx'
        #else:
        #    fname = filename + '.xlsx'
        sio = StringIO()
        wb.save(sio)
        return sio.getvalue()
